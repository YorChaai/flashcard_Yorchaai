import openpyxl

# Create sample vocabulary data
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vocabulary"

# Add headers
ws['A1'] = 'no'
ws['B1'] = 'kata'

# Add 50 sample English words
words = [
    "apple", "banana", "cherry", "date", "elderberry",
    "fig", "grape", "honeydew", "indigo", "jackfruit",
    "kiwi", "lemon", "mango", "nectarine", "orange",
    "papaya", "quince", "raspberry", "strawberry", "tangerine",
    "umbrella", "violet", "watermelon", "xylophone", "yellow",
    "zebra", "angel", "butterfly", "cloud", "diamond",
    "elephant", "flower", "garden", "harmony", "island",
    "jewel", "kingdom", "library", "mountain", "notebook",
    "ocean", "penguin", "queen", "rainbow", "sunflower",
    "tiger", "universe", "violin", "waterfall", "crystal"
]

for i, word in enumerate(words, start=1):
    ws.cell(row=i+1, column=1, value=i)
    ws.cell(row=i+1, column=2, value=word)

# Save the file
wb.save('sample_vocabulary.xlsx')
print("Created sample_vocabulary.xlsx with 50 words")
