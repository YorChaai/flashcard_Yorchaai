# -*- coding: utf-8 -*-
"""
Script untuk memproses file Excel Jepang:
1. Membaca kamus kata Jepang (dari asset/jepang.xlsx + tools/grammar_dict.json + perbaikan)
2. Membuat backup file target
3. Menyisipkan 4 kolom: Kanji, Hiragana, Katakana, Romaji di ke-4 sheet jika belum ada:
   - Main (Original)
   - Grammar (Original)
   - Main (Wuthuring)
   - Grammar (Wuthuring)
4. Mengisi data terjemahan untuk SEMUA kata yang cocok (termasuk kata grammar!)
5. Memperbarui formula Copy Prompt agar tetap sinkron ke kolom Type (H) dan CEFR (I)
6. Berjalan 100% lokal TANPA token AI
"""

import sys
import io
import os
import shutil
import time
import json
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE_DIR = r"d:\2. Organize\1. Projects\flashcard"
ASSET_DIR = os.path.join(BASE_DIR, "asset")
TOOLS_DIR = os.path.join(BASE_DIR, "tools")

JEPANG_DICT_FILE = os.path.join(ASSET_DIR, "jepang.xlsx")
GRAMMAR_DICT_FILE = os.path.join(TOOLS_DIR, "grammar_dict.json")
TARGET_FILE = os.path.join(ASSET_DIR, "flashcard_FINAL_WITH_PROMPT_V4_Jepang.xlsx")
BACKUP_FILE = os.path.join(ASSET_DIR, "flashcard_FINAL_WITH_PROMPT_V4_Jepang.xlsx.bak")

def load_all_dictionaries():
    print("[1/5] Memuat kamus kosakata & grammar Jepang...")
    dict_map = {}

    # 1. Dari jepang.xlsx (Main vocabulary)
    if os.path.exists(JEPANG_DICT_FILE):
        wb = openpyxl.load_workbook(JEPANG_DICT_FILE, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            kata = str(row[0]).strip().lower() if row[0] else ""
            if kata:
                dict_map[kata] = {
                    "kanji": row[1] if row[1] is not None else "—",
                    "hiragana": row[2] if row[2] is not None else "—",
                    "katakana": row[3] if row[3] is not None else "—",
                    "romaji": row[4] if row[4] is not None else "—",
                }

    # 2. Dari grammar_dict.json (Grammar words)
    if os.path.exists(GRAMMAR_DICT_FILE):
        with open(GRAMMAR_DICT_FILE, "r", encoding="utf-8") as f:
            grammar_data = json.load(f)
            dict_map.update(grammar_data)

    # 3. Koreksi leksikal natural
    audited_corrections = {
        'state': {'kanji': '国', 'hiragana': 'くに', 'katakana': 'ステート', 'romaji': 'kuni'},
        'software': {'kanji': '—', 'hiragana': 'そふとうぇあ', 'katakana': 'ソフトウェア', 'romaji': 'sofutowea'},
        'just': {'kanji': '—', 'hiragana': 'だけ', 'katakana': 'ジャスト', 'romaji': 'dake'},
        'used': {'kanji': '使われた', 'hiragana': 'つかわれた', 'katakana': 'ユーズド', 'romaji': 'tsukawareta'},
        'please': {'kanji': 'お願い', 'hiragana': 'どうぞ', 'katakana': 'プリーズ', 'romaji': 'dōzo'},
        'web': {'kanji': '—', 'hiragana': 'うぇぶ', 'katakana': 'ウェブ', 'romaji': 'webu'},
        'here': {'kanji': '—', 'hiragana': 'ここ', 'katakana': '—', 'romaji': 'koko'},
    }
    dict_map.update(audited_corrections)
    print(f"      Total entri kamus gabungan: {len(dict_map)} kata.")
    return dict_map

def process_sheets():
    # Cek lock file
    try:
        with open(TARGET_FILE, "r+b") as test_f:
            pass
    except PermissionError:
        print("\n[ERROR] File Excel sedang dibuka oleh program lain (misal Microsoft Excel)!")
        print("Silakan TUTUP file Excel terlebih dahulu, lalu jalankan kembali script ini.")
        sys.exit(1)

    print("[2/5] Membuat backup aman ke:", BACKUP_FILE)
    shutil.copyfile(TARGET_FILE, BACKUP_FILE)

    print("[3/5] Membuka target workbook:", TARGET_FILE)
    t0 = time.time()
    wb = openpyxl.load_workbook(TARGET_FILE)
    print(f"      Workbook berhasil dibuka dalam {time.time() - t0:.2f} detik.")

    dict_map = load_all_dictionaries()

    target_sheets = [
        ("Main (Original)", False),
        ("Grammar (Original)", False),
        ("Main (Wuthuring)", True),
        ("Grammar (Wuthuring)", True),
    ]

    print("[4/5] Memproses 4 sheet...")
    for sheet_name, has_total in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"      Peringatan: Sheet '{sheet_name}' tidak ditemukan, dilewati.")
            continue

        ws = wb[sheet_name]
        total_rows = ws.max_row
        print(f"\n      -> Memproses Sheet: '{sheet_name}' ({total_rows} baris)...")

        # Cek apakah kolom sudah disisipkan sebelumnya
        col2_val = str(ws.cell(row=1, column=2).value or '').strip()
        if col2_val.lower() != 'kanji':
            ws.insert_cols(2, amount=4)

        # Set Headers
        ws.cell(row=1, column=1, value="Kata")
        ws.cell(row=1, column=2, value="Kanji")
        ws.cell(row=1, column=3, value="Hiragana")
        ws.cell(row=1, column=4, value="Katakana")
        ws.cell(row=1, column=5, value="Romaji")
        ws.cell(row=1, column=6, value="Arti")
        ws.cell(row=1, column=7, value="IPA")
        ws.cell(row=1, column=8, value="Type")
        ws.cell(row=1, column=9, value="CEFR")
        ws.cell(row=1, column=10, value="Top")
        if has_total:
            ws.cell(row=1, column=11, value="total")
            prompt_col = 12
        else:
            prompt_col = 11
        ws.cell(row=1, column=prompt_col, value="Copy Prompt")

        # Isi data terjemahan dan perbarui formula
        matched_count = 0
        for r in range(2, total_rows + 1):
            kata_val = ws.cell(row=r, column=1).value
            kata_clean = str(kata_val).strip().lower() if kata_val else ""

            if kata_clean in dict_map:
                entry = dict_map[kata_clean]
                ws.cell(row=r, column=2, value=entry["kanji"])
                ws.cell(row=r, column=3, value=entry["hiragana"])
                ws.cell(row=r, column=4, value=entry["katakana"])
                ws.cell(row=r, column=5, value=entry["romaji"])
                matched_count += 1

            formula_val = ws.cell(row=r, column=prompt_col).value
            if formula_val and isinstance(formula_val, str) and formula_val.startswith('='):
                new_formula = f'=SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(ai!$A$2, "[KATA]", A{r}), "[TYPE]", H{r}), "[CERF]", I{r})'
                ws.cell(row=r, column=prompt_col, value=new_formula)

        print(f"         Selesai: {total_rows - 1} baris kata diproses ({matched_count} kata Jepang terisi).")

    print("\n[5/5] Menyimpan workbook yang telah diperbarui...")
    t_save = time.time()
    wb.save(TARGET_FILE)
    print(f"      File berhasil disimpan dalam {time.time() - t_save:.2f} detik ke: {TARGET_FILE}")
    print("\n=== SEMUA PROSES BERHASIL SELESAI DENGAN SUKSES (0 TOKEN AI DIGUNAKAN) ===")

if __name__ == "__main__":
    process_sheets()
