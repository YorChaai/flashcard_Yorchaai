import openpyxl

# Create sample data for different column counts

# 2 Columns: no, kata
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Data"
ws2.append(['no', 'kata'])
words_2col = [
    (1, "apple"), (2, "banana"), (3, "cherry"), (4, "date"), (5, "elderberry"),
    (6, "fig"), (7, "grape"), (8, "honeydew"), (9, "indigo"), (10, "jackfruit")
]
for row in words_2col:
    ws2.append(row)
wb2.save('sample_2columns.xlsx')
print("Created sample_2columns.xlsx")

# 3 Columns: no, kata, type
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "Data"
ws3.append(['no', 'kata', 'type'])
words_3col = [
    (1, "apple", "noun"), (2, "run", "verb"), (3, "happy", "adj"),
    (4, "quickly", "adv"), (5, "book", "noun"), (6, "jump", "verb"),
    (7, "beautiful", "adj"), (8, "slowly", "adv"), (9, "car", "noun"),
    (10, "eat", "verb")
]
for row in words_3col:
    ws3.append(row)
wb3.save('sample_3columns.xlsx')
print("Created sample_3columns.xlsx")

# 4 Columns: no, kata, type, meaning
wb4 = openpyxl.Workbook()
ws4 = wb4.active
ws4.title = "Data"
ws4.append(['no', 'kata', 'type', 'meaning'])
words_4col = [
    (1, "apple", "noun", "buah apel"), (2, "run", "verb", "berlari"),
    (3, "happy", "adj", "senang"), (4, "quickly", "adv", "dengan cepat"),
    (5, "book", "noun", "buku"), (6, "jump", "verb", "melompat"),
    (7, "beautiful", "adj", "cantik"), (8, "slowly", "adv", "dengan lambat"),
    (9, "car", "noun", "mobil"), (10, "eat", "verb", "makan")
]
for row in words_4col:
    ws4.append(row)
wb4.save('sample_4columns.xlsx')
print("Created sample_4columns.xlsx")

# 5 Columns: no, kata, type, meaning, example
wb5 = openpyxl.Workbook()
ws5 = wb5.active
ws5.title = "Data"
ws5.append(['no', 'kata', 'type', 'meaning', 'example'])
words_5col = [
    (1, "apple", "noun", "buah apel", "I eat an apple"),
    (2, "run", "verb", "berlari", "I run every morning"),
    (3, "happy", "adj", "senang", "She looks happy"),
    (4, "quickly", "adv", "dengan cepat", "He speaks quickly"),
    (5, "book", "noun", "buku", "Read this book"),
]
for row in words_5col:
    ws5.append(row)
wb5.save('sample_5columns.xlsx')
print("Created sample_5columns.xlsx")

# 6 Columns: no, kata, type, meaning, example, synonym
wb6 = openpyxl.Workbook()
ws6 = wb6.active
ws6.title = "Data"
ws6.append(['no', 'kata', 'type', 'meaning', 'example', 'synonym'])
words_6col = [
    (1, "apple", "noun", "buah apel", "I eat an apple", "fruit"),
    (2, "run", "verb", "berlari", "I run every morning", "sprint"),
    (3, "happy", "adj", "senang", "She looks happy", "joyful"),
    (4, "quickly", "adv", "dengan cepat", "He speaks quickly", "fast"),
    (5, "book", "noun", "buku", "Read this book", "novel"),
]
for row in words_6col:
    ws6.append(row)
wb6.save('sample_6columns.xlsx')
print("Created sample_6columns.xlsx")

print("\n✅ All sample files created successfully!")
